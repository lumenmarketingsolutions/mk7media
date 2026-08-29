"""
FGC WhatsApp order-entry bot.

Mary (or Kendall) WhatsApps order lines to the Lumen number (+1 623 512 6504)
in the format she already uses ("3900511 / 14$ / ghazir", names/2 pcs optional,
multiple orders per message fine). This module:

  1. Only activates for senders whitelisted in FGC_ORDER_SENDERS (comma-separated
     wa_ids, digits only). Everyone else falls through to the normal agent.
  2. Parses the message with Claude into structured orders.
  3. Creates each order in the FGC Shopify store via the draft-order flow
     (identical to manual entry: agreed price includes $3 delivery, line
     discounted, COD pending, tagged WhatsApp).
  4. Replies in the chat with a confirmation or a "couldn't read that" nudge.

Env (Railway):
  FGC_ORDER_SENDERS            e.g. "96179018107,12085910132" — REQUIRED to activate
  FGC_SHOPIFY_CLIENT_ID        Jarvis Ops app client id
  FGC_SHOPIFY_CLIENT_SECRET    Jarvis Ops app client secret
  ANTHROPIC_API_KEY            shared with the agents
Conventions (fixed): Lebanese numbers get +961, price includes $3 delivery,
default product = Migraine Relief Cap. "2 pcs" = quantity 2 at that total.
"""

import os
import json
import time
import threading
import urllib.request

SHOP = "https://hd8wtv-ck.myshopify.com"
API_VERSION = "2025-07"
CAP_VARIANT = 46831330427079
# Fallback prices only — the live price is fetched from Shopify per order (see
# _live_unit_price). NEVER trust these for discount math: when the store gets
# repriced and this table lags, the discount lands on the wrong base price
# (08.25-08.27 bug: cap repriced 19.99 -> 12.00, orders entered at 8.51 instead
# of the agreed 16.50).
VARIANTS = {  # product keyword -> (variant_id, fallback_unit_price)
    "cap": (46831330427079, 12.00),
    "patches": (46831330721991, 12.00),
    "strips": (46831330885831, 12.00),
}
# Human names for the confirmation reply — Mary should see WHAT was ordered, not
# just a total, so a wrong-product entry is obvious the moment it is made.
PRODUCT_NAMES = {
    "cap": "Migraine Cap",
    "patches": "Pimple Patches",
    "strips": "Whitening Strips",
}

FGC_ORDER_SENDERS = {
    "".join(ch for ch in s if ch.isdigit())
    for s in os.environ.get("FGC_ORDER_SENDERS", "").split(",") if s.strip()
}
SHOPIFY_CLIENT_ID = os.environ.get("FGC_SHOPIFY_CLIENT_ID", "")
SHOPIFY_CLIENT_SECRET = os.environ.get("FGC_SHOPIFY_CLIENT_SECRET", "")
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
PARSE_MODEL = os.environ.get("FGC_ORDER_PARSE_MODEL", "claude-sonnet-5")

_token_cache = {"token": None, "expires": 0}
_last_errors = []  # ring buffer of recent failures for the /fgc-orders/health endpoint


def _record_error(where, err):
    _last_errors.append({"at": time.strftime("%Y-%m-%d %H:%M:%S"), "where": where, "error": str(err)[:300]})
    del _last_errors[:-10]
    print(f"[fgc-orders] {where}: {err}")


def _xlsx_available():
    try:
        import openpyxl  # noqa: F401
        return True
    except ImportError:
        return False


def health():
    """Status dict for the admin debug endpoint — no secrets, live checks."""
    out = {
        "senders_whitelisted": sorted(FGC_ORDER_SENDERS),
        "shopify_creds_present": bool(SHOPIFY_CLIENT_ID and SHOPIFY_CLIENT_SECRET),
        "anthropic_key_present": bool(ANTHROPIC_API_KEY),
        "whatsapp_token_present": bool(os.environ.get("WHATSAPP_ACCESS_TOKEN")),
        "sheet_import": {"csv": True, "xlsx": _xlsx_available()},
        "recent_errors": list(_last_errors),
    }
    try:
        _shopify_token()
        shop = _shopify("GET", "/shop.json?fields=name")
        out["shopify_connection"] = "OK - " + str((shop.get("shop") or {}).get("name", shop))
    except Exception as e:
        out["shopify_connection"] = "FAILED - " + str(e)[:200]
    return out


def is_order_sender(wa_id):
    return bool(FGC_ORDER_SENDERS) and "".join(ch for ch in str(wa_id) if ch.isdigit()) in FGC_ORDER_SENDERS


def _shopify_token():
    if _token_cache["token"] and time.time() < _token_cache["expires"] - 300:
        return _token_cache["token"]
    body = json.dumps({"grant_type": "client_credentials",
                       "client_id": SHOPIFY_CLIENT_ID,
                       "client_secret": SHOPIFY_CLIENT_SECRET}).encode()
    req = urllib.request.Request(SHOP + "/admin/oauth/access_token", body, method="POST")
    req.add_header("Content-Type", "application/json")
    with urllib.request.urlopen(req, timeout=20) as r:
        d = json.loads(r.read())
    _token_cache["token"] = d["access_token"]
    _token_cache["expires"] = time.time() + int(d.get("expires_in", 86400))
    return _token_cache["token"]


def _shopify(method, path, body=None):
    req = urllib.request.Request(SHOP + f"/admin/api/{API_VERSION}" + path, method=method)
    req.add_header("Content-Type", "application/json")
    req.add_header("X-Shopify-Access-Token", _shopify_token())
    data = json.dumps(body).encode() if body is not None else b""
    req.add_header("Content-Length", str(len(data)))
    try:
        with urllib.request.urlopen(req, data if method in ("POST", "PUT") else None, timeout=30) as r:
            return json.loads(r.read() or b"{}")
    except urllib.error.HTTPError as e:
        return {"_error": e.code, "body": (e.read() or b"")[:300].decode("utf-8", "replace")}


PARSE_PROMPT = """You parse WhatsApp order messages for a Lebanese cash-on-delivery store. \
Messages contain one or more orders, typically "phone / price / city" with optional name, \
"2 pcs"/"2 pieces"/"2 boxes" (quantity), address details, or product hints (cap/patches/strips — \
default is cap). Reply ONLY with JSON: {"orders": [{"phone": "...", "price": 16, \
"quantity": 1, "city": "...", "name": "", "address": "", "product": "cap"}], \
"not_orders": "text that wasn't parseable as an order, or empty string"}. \
Rules: phone exactly as written (keep leading zeros / country codes). price = the total \
number they wrote (it includes delivery). quantity from "2 pcs" style notes, else 1. \
If the message contains no orders at all, return {"orders": [], "not_orders": "<the text>"}."""


def _parse_orders(text):
    body = json.dumps({
        "model": PARSE_MODEL, "max_tokens": 2000,
        "system": PARSE_PROMPT,
        "messages": [{"role": "user", "content": text}],
    }).encode()
    req = urllib.request.Request("https://api.anthropic.com/v1/messages", body, method="POST")
    req.add_header("Content-Type", "application/json")
    req.add_header("x-api-key", ANTHROPIC_API_KEY)
    req.add_header("anthropic-version", "2023-06-01")
    with urllib.request.urlopen(req, timeout=60) as r:
        d = json.loads(r.read())
    raw = "".join(b.get("text", "") for b in d.get("content", []))
    raw = raw[raw.find("{"): raw.rfind("}") + 1]
    return json.loads(raw)


import re

# Accepts: "3900511 / 14$ / ghazir", "03-900-511 - 14 - ghazir", "3900511 14$ ghazir",
# "+961 3 900 511 | 16.5$ | Beirut 2 pcs", optional leading name, optional product word.
_SEP = r"\s*(?:/|-|\||,|\s)\s*"
_LINE_RE = re.compile(
    r"^\s*(?P<name>[A-Za-z\u0600-\u06FF][A-Za-z\u0600-\u06FF .']{1,30}?)?\s*"
    r"(?P<phone>\+?\d[\d \-]{5,16}\d)"
    + _SEP +
    r"\$?\s*(?P<price>\d{1,3}(?:[.,]\d{1,2})?)\s*(?:\$|usd|dollars?)?"
    r"(?:" + _SEP + r"(?P<rest>.+))?\s*$",
    re.I,
)
_QTY_RE = re.compile(r"(?:(?P<a>\d)\s*(?:pcs?|pieces?|pc|boxes|box|x)\b|\bx\s*(?P<b>\d)\b)", re.I)
_PRODUCT_WORDS = {"patches": "patches", "patch": "patches", "pimple": "patches",
                  "strips": "strips", "strip": "strips", "whitening": "strips", "cap": "cap", "caps": "cap"}


def _regex_parse(text):
    """Deterministic parser for the 'number / price / city' format — runs FIRST so
    orders never depend on the Claude API (credits, outages)."""
    orders, leftovers = [], []
    for line in text.splitlines():
        line = line.strip().strip("•-* ")
        if not line:
            continue
        m = _LINE_RE.match(line)
        if not m:
            leftovers.append(line)
            continue
        rest = (m.group("rest") or "").strip()
        qm = _QTY_RE.search(rest)
        qty = int(qm.group("a") or qm.group("b")) if qm else 1
        product = "cap"
        for w, prod in _PRODUCT_WORDS.items():
            if re.search(r"\b" + w + r"\b", rest, re.I):
                product = prod
                rest = re.sub(r"\b" + w + r"\b", "", rest, flags=re.I)
                break
        city = _QTY_RE.sub("", rest).strip(" /|,-") or "Lebanon"
        phone = re.sub(r"[\s\-]", "", m.group("phone"))
        price = float(m.group("price").replace(",", "."))
        orders.append({"phone": phone, "price": price, "quantity": max(1, qty), "city": city,
                       "name": (m.group("name") or "").strip(), "address": "", "product": product})
    return {"orders": orders, "not_orders": "\n".join(leftovers)}


_price_cache = {}  # variant_id -> (price, fetched_at)


def _live_unit_price(variant_id, fallback):
    """Current store price for a variant, cached 1h. Falls back to the static
    table only if Shopify can't be reached (order still gets total-verified)."""
    hit = _price_cache.get(variant_id)
    if hit and time.time() - hit[1] < 3600:
        return hit[0]
    try:
        v = _shopify("GET", f"/variants/{variant_id}.json?fields=price")
        price = float(v["variant"]["price"])
        _price_cache[variant_id] = (price, time.time())
        return price
    except Exception as e:
        _record_error("live_price", f"variant {variant_id}: {e}")
        return hit[0] if hit else fallback


# ---------------------------------------------------------------- duplicates
# A sheet can be sent twice, or a corrected sheet can overlap an earlier one, so
# every import checks the store first. The key is the customer's phone number:
# these are one-off COD orders, so the same number appearing again is almost
# always a re-import rather than a genuine second purchase.
_orders_cache = {"index": None, "at": 0}


def _norm_digits(raw):
    d = "".join(ch for ch in str(raw or "") if ch.isdigit())
    if d.startswith("961"):
        d = d[3:]
    return d.lstrip("0")


def _existing_orders_index(max_pages=6):
    """{phone_digits: [{name, total, product}]} for recent store orders, cached 5min."""
    if _orders_cache["index"] is not None and time.time() - _orders_cache["at"] < 300:
        return _orders_cache["index"]
    idx = {}
    path = ("/orders.json?status=any&limit=250"
            "&fields=id,name,total_price,shipping_address,line_items,created_at")
    try:
        for _ in range(max_pages):
            d = _shopify("GET", path)
            orders = d.get("orders") or []
            if not orders:
                break
            for o in orders:
                ph = _norm_digits(((o.get("shipping_address") or {}).get("phone")))
                if not ph:
                    continue
                items = ", ".join((li.get("title") or "") for li in (o.get("line_items") or [])[:2])
                idx.setdefault(ph, []).append({
                    "name": o.get("name"), "total": float(o.get("total_price") or 0),
                    "product": items, "created": (o.get("created_at") or "")[:10]})
            if len(orders) < 250:
                break
            last = orders[-1]["id"]
            path = ("/orders.json?status=any&limit=250&since_id=%s"
                    "&fields=id,name,total_price,shipping_address,line_items,created_at" % last)
        _orders_cache["index"] = idx
        _orders_cache["at"] = time.time()
    except Exception as e:
        _record_error("orders_index", e)
        return _orders_cache["index"] or {}
    return idx


def _duplicate_of(o, index):
    """Returns (existing_order_dict, exact_bool) if this order looks already entered."""
    hits = index.get(_norm_digits(o.get("phone")))
    if not hits:
        return None, False
    price = float(o.get("price") or 0)
    for h in hits:
        if abs(h["total"] - price) <= 0.01:
            return h, True          # same number, same total -> re-import
    return hits[0], False           # same number, different total -> flag, still enter


# ------------------------------------------------------------------ media/sheets
GRAPH = "https://graph.facebook.com/v21.0"


def _download_media(media_id):
    """WhatsApp media is two hops: metadata (gives a signed url), then the bytes."""
    import requests
    tok = os.environ.get("WHATSAPP_ACCESS_TOKEN", "")
    if not tok:
        raise RuntimeError("WHATSAPP_ACCESS_TOKEN not set")
    h = {"Authorization": "Bearer " + tok}
    meta = requests.get(f"{GRAPH}/{media_id}", headers=h, timeout=30).json()
    url = meta.get("url")
    if not url:
        raise RuntimeError(f"no media url: {str(meta)[:200]}")
    r = requests.get(url, headers=h, timeout=90)
    r.raise_for_status()
    return r.content, (meta.get("mime_type") or "")


_COL_HINTS = {
    "phone": ("phone", "number", "mobile", "tel", "whatsapp", "contact", "num"),
    "price": ("price", "total", "amount", "cost", "paid", "value", "$"),
    "city":  ("city", "area", "location", "region", "town", "address", "delivery"),
    "name":  ("name", "customer", "client", "full name"),
    "quantity": ("qty", "quantity", "pcs", "pieces", "count"),
    "product": ("product", "item", "sku", "type"),
}


def _map_columns(header):
    """header list -> {field: column index}. Unrecognised columns are ignored."""
    out = {}
    for i, cell in enumerate(header):
        c = str(cell or "").strip().lower()
        if not c:
            continue
        for field, hints in _COL_HINTS.items():
            if field in out:
                continue
            if any(h in c for h in hints):
                out[field] = i
                break
    return out


def _rows_from_bytes(data, filename, mime):
    """CSV or XLSX bytes -> list of row lists."""
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")) or "spreadsheetml" in (mime or ""):
        import io
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("xlsx support not installed on the server (openpyxl)")
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        return [list(r) for r in ws.iter_rows(values_only=True)]
    import csv, io as _io
    txt = data.decode("utf-8-sig", "replace")
    # Try every common delimiter and keep the parse that actually yields a usable
    # table. Sniffer alone is unreliable on hand-made sheets (mixed separators,
    # semicolons from Excel in some locales, a stray comma inside a name).
    best, best_score = None, -1
    for delim in (",", ";", "\t", "|"):
        try:
            rows = [r for r in csv.reader(_io.StringIO(txt), delimiter=delim)]
        except Exception:
            continue
        if not rows:
            continue
        widths = [len(r) for r in rows if any(str(c or "").strip() for c in r)]
        if not widths:
            continue
        width = max(set(widths), key=widths.count)      # most common row width
        score = width * 10
        for r in rows[:5]:                              # a mappable header is worth more
            m = _map_columns(r)
            if "phone" in m and "price" in m:
                score += 100
                break
        if width < 2:
            score -= 50
        if score > best_score:
            best, best_score = rows, score
    return best or [r for r in csv.reader(_io.StringIO(txt))]


def _orders_from_sheet(data, filename, mime):
    """Parse a sheet into the same order dicts the text parser produces.
    Returns (orders, skipped_rows)."""
    rows = [r for r in _rows_from_bytes(data, filename, mime) if any(str(c or "").strip() for c in r)]
    if not rows:
        return [], ["sheet is empty"]
    cols, start = {}, 0
    for i, r in enumerate(rows[:5]):          # header may not be the first row
        m = _map_columns(r)
        if "phone" in m and "price" in m:
            cols, start = m, i + 1
            break
    orders, skipped = [], []
    for r in rows[start:]:
        get = lambda f: (str(r[cols[f]]).strip() if f in cols and cols[f] < len(r) and r[cols[f]] is not None else "")
        if cols:
            phone, price_raw = get("phone"), get("price")
            name, city = get("name"), get("city")
            qty_raw, prod_raw = get("quantity"), get("product")
        else:                                  # no usable header — fall back to the text parser
            line = " / ".join(str(c).strip() for c in r if str(c or "").strip())
            p = _regex_parse(line)
            if p["orders"]:
                orders.extend(p["orders"])
            elif line:
                skipped.append(line[:60])
            continue
        price_digits = re.sub(r"[^\d.]", "", price_raw.replace(",", "."))
        if not phone or not price_digits:
            skipped.append((" / ".join(x for x in (name, phone, price_raw, city) if x))[:60] or "blank row")
            continue
        product = "cap"
        blob = f"{prod_raw} {city}".lower()
        for w, prod in _PRODUCT_WORDS.items():
            if re.search(r"\b" + w + r"\b", blob):
                product = prod
                break
        try:
            qty = int(re.sub(r"[^\d]", "", qty_raw) or 1)
        except ValueError:
            qty = 1
        orders.append({"phone": phone, "price": float(price_digits), "quantity": max(1, qty),
                       "city": city or "Lebanon", "name": name, "address": "", "product": product})
    return orders, skipped


def _norm_phone(raw):
    digits = "".join(ch for ch in str(raw) if ch.isdigit())
    if str(raw).strip().startswith("+") and not digits.startswith("961"):
        return "+" + digits          # explicit foreign number, keep as written
    if digits.startswith("961"):
        return "+" + digits
    if digits.startswith("0"):
        digits = digits[1:]
    return "+961" + digits


def _enter_order(o):
    product = (o.get("product") or "cap").lower()
    variant_id, fallback = VARIANTS.get(product, VARIANTS["cap"])
    unit = _live_unit_price(variant_id, fallback)
    qty = max(1, int(o.get("quantity") or 1))
    price = float(o["price"])
    goods = round(price - 3, 2)
    disc_per_unit = max(0.0, round((unit * qty - goods) / qty, 2))
    # If the agreed price is above list, no discount applies; put the difference on
    # the delivery line so the order total always equals what was agreed.
    shipping = round(price - (unit * qty - disc_per_unit * qty), 2)
    name = (o.get("name") or "").strip()
    first, last = (name.split(" ", 1) + [""])[:2] if name else ("WhatsApp", "".join(ch for ch in str(o["phone"]) if ch.isdigit()))
    addr = (o.get("address") or "").strip() or o.get("city") or "Lebanon"
    note = f"Auto-entered from WhatsApp. Agreed {price:g}$ incl 3$ delivery."
    body = {"draft_order": {
        "line_items": [{"variant_id": variant_id, "quantity": qty,
                        "applied_discount": {"description": "WhatsApp agreed price",
                                             "value_type": "fixed_amount",
                                             "value": str(disc_per_unit), "amount": str(disc_per_unit)}}],
        "shipping_address": {"first_name": first, "last_name": last or first,
                             "address1": addr, "city": o.get("city") or "Lebanon",
                             "country": "Lebanon", "phone": _norm_phone(o["phone"])},
        "shipping_line": {"title": "Delivery", "price": f"{shipping:.2f}"},
        "tags": "WhatsApp, auto-entry",
        "note": note}}
    d = _shopify("POST", "/draft_orders.json", body)
    draft = d.get("draft_order")
    if not draft:
        return None, f"draft failed: {d.get('body', d)}"
    # HARD RULE: the order total must equal the agreed price before it becomes a
    # real order. If Shopify's computed total disagrees (repriced product, math
    # drift), the draft is deleted and nothing is entered.
    draft_total = float(draft.get("total_price") or 0)
    if abs(draft_total - price) > 0.01:
        _shopify("DELETE", f"/draft_orders/{draft['id']}.json")
        _record_error("total_mismatch",
                      f"draft total {draft_total:.2f} != agreed {price:g} (variant {variant_id}, qty {qty})")
        return None, (f"store total came out ${draft_total:.2f} but agreed price is ${price:g} — "
                      "NOT entered (store prices may have changed). Tell Kendall.")
    done = _shopify("PUT", f"/draft_orders/{draft['id']}/complete.json?payment_pending=true")
    order_id = (done.get("draft_order") or {}).get("order_id")
    if not order_id:
        return None, f"complete failed: {done.get('body', done)}"
    order = _shopify("GET", f"/orders/{order_id}.json?fields=name,total_price").get("order", {})
    order["_product"] = PRODUCT_NAMES.get(product, product.title())
    order["_qty"] = qty
    return order, None


def _confirm_line(order, o):
    """✅ #1201 · Migraine Cap ×2 · $16.50 · Ghazir — Mary sees the item, the money
    and the delivery area, so a wrong product is obvious immediately."""
    qty = order.get("_qty") or 1
    item = order.get("_product") or "Item"
    if qty > 1:
        item += f" \u00d7{qty}"
    city = (o.get("city") or "").strip()
    bits = [f"\u2705 {order.get('name')}", item, f"${order.get('total_price')}"]
    if city:
        bits.append(city)
    return " \u00b7 ".join(bits)


def handle_order_message(wa_id, text, send_text):
    """Entry point — called from the webhook route for whitelisted senders.
    Runs async so the webhook can ack immediately."""
    def work():
        try:
            if not (SHOPIFY_CLIENT_ID and SHOPIFY_CLIENT_SECRET):
                send_text(wa_id, "Order bot isn't configured yet (missing store credentials). Tell Kendall.")
                return
            try:
                _shopify_token()
            except Exception as e:
                _record_error("shopify_token", e)
                send_text(wa_id, "Order bot can't reach the store (credentials rejected). Tell Kendall to check the Shopify vars.")
                return
            parsed = _regex_parse(text)
            if not parsed.get("orders"):
                # Standard format didn't match -> let Claude try the messy version.
                try:
                    parsed = _parse_orders(text)
                except Exception as e:
                    _record_error("parse", f"{e} | text={text[:120]!r}")
                    send_text(wa_id, "Order bot couldn't read that. Use: number / price / city (one order per line).")
                    return
            orders = parsed.get("orders") or []
            if not orders:
                send_text(wa_id, "I couldn't read an order in that. Format: number / price / city (name and \"2 pcs\" optional).")
                return
            lines, fails = [], []
            for o in orders:
                order, err = _enter_order(o)
                if order:
                    lines.append(_confirm_line(order, o))
                else:
                    fails.append(f"⚠️ {o.get('phone')} / {o.get('price')}$ — {err}")
            reply = "\n".join(lines + fails)
            if len(orders) > 1:
                reply = f"Entered {len(lines)}/{len(orders)} orders:\n" + reply
            send_text(wa_id, reply[:3900])
        except Exception as e:
            _record_error("entry", e)
            try:
                send_text(wa_id, "Something broke entering that order — Kendall's been notified, try once more or send it to him.")
            except Exception:
                pass
    threading.Thread(target=work, daemon=True).start()


def handle_order_sheet(wa_id, media_id, filename, caption, send_text):
    """Mary sends a CSV/XLSX to the order number and every row becomes an order.

    Duplicate-safe: before entering anything the store's recent orders are indexed
    by phone number. A row whose number already has an order at the same total is
    SKIPPED (that is a re-sent or overlapping sheet). A row whose number exists at
    a different total is still entered but flagged in the reply, so a genuine
    second purchase goes through while a mistake is visible.

    Put "force" in the caption to enter every row regardless.
    """
    def work():
        try:
            if not (SHOPIFY_CLIENT_ID and SHOPIFY_CLIENT_SECRET):
                send_text(wa_id, "Order bot isn't configured yet (missing store credentials). Tell Kendall.")
                return
            send_text(wa_id, f"Got the sheet ({filename or 'file'}) — reading it now, one moment.")
            try:
                data, mime = _download_media(media_id)
            except Exception as e:
                _record_error("media", e)
                send_text(wa_id, "I couldn't download that file from WhatsApp. Try sending it again as a document.")
                return
            try:
                orders, skipped = _orders_from_sheet(data, filename, mime)
            except Exception as e:
                _record_error("sheet_parse", f"{e} | {filename}")
                send_text(wa_id, f"I couldn't read that sheet ({e}). It needs a phone column and a price column — "
                                 "CSV or Excel both work.")
                return
            if not orders:
                send_text(wa_id, "I read the sheet but found no orders in it. It needs a column with the phone "
                                 "number and one with the price. City, name, quantity and product are optional.")
                return

            force = "force" in (caption or "").lower()
            index = {} if force else _existing_orders_index()
            entered, dups, flagged, fails = [], [], [], []
            for o in orders:
                if not force:
                    hit, exact = _duplicate_of(o, index)
                    if hit and exact:
                        dups.append(f"{o.get('phone')} — already entered as {hit['name']} (${hit['total']:.2f})")
                        continue
                    if hit:
                        flagged.append(f"{o.get('phone')} — also has {hit['name']} (${hit['total']:.2f})")
                order, err = _enter_order(o)
                if order:
                    entered.append(_confirm_line(order, o))
                    ph = _norm_digits(o.get("phone"))
                    index.setdefault(ph, []).append({"name": order.get("name"),
                                                     "total": float(order.get("total_price") or 0),
                                                     "product": order.get("_product", ""), "created": ""})
                else:
                    fails.append(f"⚠️ {o.get('phone')} / {o.get('price')}$ — {err}")

            head = f"Sheet done — {len(entered)} entered"
            if dups:
                head += f", {len(dups)} skipped as already in the store"
            if fails:
                head += f", {len(fails)} failed"
            parts = [head + "."]
            if entered:
                parts.append("\n".join(entered))
            if dups:
                parts.append("Already in the store (not entered again):\n" + "\n".join("• " + d for d in dups))
            if flagged:
                parts.append("Entered, but this number already had an order — check these:\n"
                             + "\n".join("• " + d for d in flagged))
            if fails:
                parts.append("\n".join(fails))
            if skipped:
                parts.append(f"{len(skipped)} row(s) I couldn't read:\n" + "\n".join("• " + r for r in skipped[:8]))
            if dups and not force:
                parts.append('If those really are new orders, send the sheet again with "force" in the caption.')
            send_text(wa_id, "\n\n".join(parts)[:3900])
        except Exception as e:
            _record_error("sheet", e)
            try:
                send_text(wa_id, "Something broke reading that sheet — Kendall's been notified.")
            except Exception:
                pass
    threading.Thread(target=work, daemon=True).start()
